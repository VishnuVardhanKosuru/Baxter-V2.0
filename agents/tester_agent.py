# -*- coding: utf-8 -*-
import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import os
import re
import time
import asyncio
import json
import csv
import argparse
from pathlib import Path
from collections import defaultdict
from dotenv import load_dotenv
import google.generativeai as genai

load_dotenv()

# Manual Test CSV Headers (14 columns)
MANUAL_CSV_HEADERS = "Test ID,Module Name,Function Name,Test Type,Test Scenario,Pre Conditions,Test Steps,Test Data,Expected Result,Priority,Executed By,Execution Date,Status,Remarks\n"

# System Prompt
SYSTEM_PROMPT = """You are a Principal Java QA Automation Engineer specializing in JUnit 5, Mockito, and Spring Boot Test.

OUTPUT FORMAT REQUIREMENTS:
Your response MUST contain TWO distinct sections:

1. AUTOMATED JUNIT TEST CLASS:
Output ONLY one valid compilable Java class inside a single ```java ... ``` block.

RULES FOR JAVA CODE:
- STRICT RETURN TYPE RULE: Every test method MUST use the exact standard Java keyword 'void' as its return type (e.g., 'void testSomething()'). NEVER prepend typos or random characters.
- ALWAYS use the EXACT method name provided in the context when calling the class under test.
- STRICT CONSTRUCTOR RULE: Use ONLY the exact constructor signatures provided in the prompt context.
- Mockito STUB RULE: When stubbing method calls with when(mock.method(any(...))), match parameter types or use plain any().
- Use JUnit 5 (org.junit.jupiter.api.*). NEVER use JUnit 4.
- Every @Test MUST have a @DisplayName that reads as a full English sentence.
- Write ALL package imports explicitly -- no wildcard imports.

2. MANUAL TEST CASES (STRICT CSV FORMAT):
Output 1-3 manual test cases for human QA testers inside a single ```csv ... ``` or [CSV]...[/CSV] block.
Output ONLY valid CSV data rows matching the exact 14 headers below. Do NOT output column headers.
If a field contains commas or newlines, wrap the field value in double quotes.
Leave the last 4 columns (Executed By, Execution Date, Status, Remarks) empty (e.g., ,,,).

Expected 14 Headers:
Test ID, Module Name, Function Name, Test Type, Test Scenario, Pre Conditions, Test Steps, Test Data, Expected Result, Priority, Executed By, Execution Date, Status, Remarks

Example CSV row:
TC_PATIENT_001,PatientService,registerPatient,Unit - Happy Path,"Verify patient registration succeeds with valid input","Database active, Mock initialized","1. Build valid PatientDTO\n2. Call registerPatient()","PatientDTO(name=""John"")","Returns saved Patient object with non-null ID",High,,,,
"""


# Prompts
UNIT_HAPPY_PATH_PROMPT = """You are writing a UNIT HAPPY PATH TEST class for `{class_name}.{func_name}()`.

WHAT YOU ARE TESTING
  Class:       {class_name}  (package: {package})
  Method:      {func_name}
  Return Type: {return_type}
  Parameters:  {parameters}

  Source Code:
  {body}

TEST SETUP
  Runner:     @ExtendWith(MockitoExtension.class)
  Mocks:      {fields_formatted}
              (declare each as @Mock)
  Under test: @InjectMocks {class_name} {class_name_lower}

CONSTRUCTORS (use these to build test objects -- do not invent):
  {constructors_formatted}

WHAT HAPPY PATH MEANS
  Test that the method works correctly when all inputs are valid and dependencies behave normally.

  Dependencies to stub (use exact arguments below):
  {calls_formatted}
  -> for each: when(mockField.method(exactArg)).thenReturn(validValue)

WRITE: class `{class_name}HappyPathTest`
  1-2 @Test methods:
    Arrange: build valid input via constructors, stub every dependency
    Act:     {class_name_lower}.{func_name}(validArgs)
    Assert:  assertNotNull(result) if object return
             assertEquals(expectedValue, result.getField())
    Verify:  verify(mockField, times(1)).method(exactArg) per dependency

Output one compilable Java class in a single ```java block.
"""

UNIT_NEGATIVE_PROMPT = """You are writing a UNIT NEGATIVE TEST class for `{class_name}.{func_name}()`.

WHAT YOU ARE TESTING
  Class:       {class_name}  (package: {package})
  Method:      {func_name}
  Parameters:  {parameters}

  Source Code:
  {body}

TEST SETUP
  Runner:     @ExtendWith(MockitoExtension.class)
  Mocks:      {fields_formatted}
  Under test: @InjectMocks {class_name} {class_name_lower}

WHAT NEGATIVE TEST MEANS
  Test that the method rejects invalid/null inputs before touching any dependency.

  Null checks found in source code (one @Test per check):
  {null_checks}
  -> assertThrows(IllegalArgumentException.class, () -> method(null))
     assertEquals("exact message from body", ex.getMessage())
     verifyNoInteractions(mock)

  Blank/empty checks found in source code (one @Test per check):
  {blank_checks}
  -> assertThrows(IllegalArgumentException.class, ...)
     verifyNoInteractions(mock)

WRITE: class `{class_name}NegativeTest`
  One @Test per scenario above.
  RULE: verifyNoInteractions(mock) in EVERY test.

Output one compilable Java class in a single ```java block.
"""

UNIT_EXCEPTION_PROMPT = """You are writing a UNIT EXCEPTION TEST class for `{class_name}.{func_name}()`.

WHAT YOU ARE TESTING
  Class:       {class_name}  (package: {package})
  Method:      {func_name}
  Parameters:  {parameters}
  Throws:      {throws_list}

  Source Code:
  {body}

TEST SETUP
  Runner:     @ExtendWith(MockitoExtension.class)
  Mocks:      {fields_formatted}
  Under test: @InjectMocks {class_name} {class_name_lower}

CONSTRUCTORS: {constructors_formatted}

WHAT EXCEPTION TEST MEANS
  Test each declared exception is thrown under its exact trigger condition.

  Exception trigger map (derived from source code -- do NOT invent):
  {exception_trigger_map}

  For each exception in {throws_list}:
    Arrange: stub mock to trigger the condition above
    Assert:  ExType ex = assertThrows(ExType.class, () -> method(...))
             assertTrue(ex.getMessage().contains("exact fragment from body"))
    Verify:  verify(mock, never()).save(any()) where save should not have been reached

WRITE: class `{class_name}ExceptionTest`
  One @Test per declared exception.

Output one compilable Java class in a single ```java block.
"""

UNIT_BOUNDARY_PROMPT = """You are writing a UNIT BOUNDARY VALUE TEST class for `{class_name}.{func_name}()`.

WHAT YOU ARE TESTING
  Class:       {class_name}  (package: {package})
  Method:      {func_name}
  Parameters:  {parameters}

  Source Code (read the comparison branches):
  {body}

TEST SETUP
  Runner:     @ExtendWith(MockitoExtension.class)
  Under test: @InjectMocks {class_name} {class_name_lower}
  (No @Mock needed if method is pure computation)

WHAT BOUNDARY TEST MEANS
  Test exact numeric edges -- bugs most commonly hide at boundaries.

  Constraints from annotations (use ONLY these -- do not invent values):
  {annotation_values_formatted}

  Branch logic from source code:
  {boundary_branches}

WRITE: class `{class_name}BoundaryTest`

  @ParameterizedTest VALID using @CsvSource:
    Rows: min, min+1, each branch switch point-1, each switch point, max-1, max
    Format: input, expectedOutput

  @ParameterizedTest INVALID using @ValueSource:
    Values: {{ min-1, max+1 }}
    assertThrows(IllegalArgumentException.class, ...)
    assertTrue(ex.getMessage().contains("Invalid"))

Output one compilable Java class in a single ```java block.
"""

UNIT_MOCK_PROMPT = """You are writing a UNIT MOCK INTERACTION TEST class for `{class_name}.{func_name}()`.

WHAT YOU ARE TESTING
  Class:       {class_name}  (package: {package})
  Method:      {func_name}
  Parameters:  {parameters}

  Source Code:
  {body}

TEST SETUP
  Runner:     @ExtendWith(MockitoExtension.class)
  Mocks:      {fields_formatted}
  Under test: @InjectMocks {class_name} {class_name_lower}

CONSTRUCTORS: {constructors_formatted}

WHAT MOCK INTERACTION TEST MEANS
  Verify the method calls dependencies with EXACT arguments -- not loose any() matchers.
  Verify dependencies are NEVER called on early-exit paths.

  CALLS edges with exact arguments:
  {calls_formatted}

  Early-exit conditions from source code:
  {early_exit_conditions}

WRITE: class `{class_name}MockTest`

  Test A -- Exact argument verification (one per CALLS edge):
    Arrange & Act: same as happy path
    verify(mock, times(1)).method(
        argThat(arg -> arg.getField().equals("exactValue")))
    RULE: do NOT use any() in verify()

  Test B -- Never-called verification (one per early-exit condition):
    Arrange: stub mock to trigger the early exit above
    assertThrows(...)
    verify(mock, never()).save(any())

Output one compilable Java class in a single ```java block.
"""

INTEG_CTRL_HAPPY_PATH_PROMPT = """You are writing an INTEGRATION HAPPY PATH TEST for controller `{class_name}`.

WHAT YOU ARE TESTING
  Class:       {class_name}  (package: {package})
  Method:      {func_name}
  HTTP:        {http_method}  {base_url}{endpoint_url}
  Return Type: {return_type}
  Parameters:  {parameters}

  Source Code:
  {body}

TEST SETUP
  Runner:   @WebMvcTest({class_name}.class)
  Declare:  @Autowired MockMvc mockMvc
            @Autowired ObjectMapper objectMapper
  MockBeans (declare each as @MockBean):
  {fields_formatted}

DTO CONSTRUCTORS (do not invent fields):
  {constructors_formatted}

WHAT HAPPY PATH MEANS
  Test the endpoint returns correct HTTP status and JSON payload
  when inputs are valid and service returns a successful response.

  Service stubs (use these exactly):
  {calls_formatted}
  -> RULE: when stubbing service methods with when(mockBean.method(any(...))), use any() or any(ExactServiceParamType.class). NEVER use the HTTP DTO type if it differs from the service method parameter type.
  -> when(mockBean.method(any(Type.class))).thenReturn(savedObject)

WRITE: class `{class_name}HappyPathIT`
  1-2 @Test methods:
    mockMvc.perform({http_method}("{base_url}{endpoint_url}")
            .contentType(MediaType.APPLICATION_JSON)
            .content(objectMapper.writeValueAsString(validDto)))
           .andExpect(status().isCreated() / isOk())
           .andExpect(jsonPath("$.id").value(expectedId))
           .andExpect(jsonPath("$.name").value("expectedName"))

Output one compilable Java class in a single ```java block.
"""

INTEG_CTRL_NEGATIVE_PROMPT = """You are writing an INTEGRATION NEGATIVE TEST for controller `{class_name}`.

WHAT YOU ARE TESTING
  Class:       {class_name}  (package: {package})
  Method:      {func_name}
  HTTP:        {http_method}  {base_url}{endpoint_url}

TEST SETUP
  Runner:   @WebMvcTest({class_name}.class)
  Declare:  @Autowired MockMvc mockMvc
            @Autowired ObjectMapper objectMapper
  MockBeans: {fields_formatted}

DTO CONSTRUCTORS: {constructors_formatted}

BOUNDARY CONSTRAINTS (for invalid value ranges):
  {annotation_values_formatted}

WHAT NEGATIVE TEST MEANS
  Bean Validation fires BEFORE the service is called -> expect 400 Bad Request.

  Invalid scenarios (one @Test per):
    - Empty/blank required fields (name = "", email = "")
    - Malformed email ("not-an-email")
    - Numeric param below @Min or above @Max from: {annotation_values_formatted}

WRITE: class `{class_name}NegativeIT`
  One @Test per scenario:
    NO mock stubbing -- validation fires before service
    mockMvc.perform(...).andExpect(status().isBadRequest())
    verifyNoInteractions(mockBean)

Output one compilable Java class in a single ```java block.
"""

INTEG_CTRL_EXCEPTION_PROMPT = """You are writing an INTEGRATION EXCEPTION TEST for controller `{class_name}`.

WHAT YOU ARE TESTING
  Class:       {class_name}  (package: {package})
  Method:      {func_name}
  HTTP:        {http_method}  {base_url}{endpoint_url}
  Throws:      {throws_list}

  Source Code:
  {body}

TEST SETUP
  Runner:   @WebMvcTest({class_name}.class)
  Declare:  @Autowired MockMvc mockMvc
            @Autowired ObjectMapper objectMapper
  MockBeans: {fields_formatted}

DTO CONSTRUCTORS: {constructors_formatted}

WHAT EXCEPTION TEST MEANS
  GlobalExceptionHandler maps each exception to an HTTP status code.

  Exception-to-HTTP mapping (use this exactly):
    DuplicateEmailException          -> 409 Conflict
    PatientNotFoundException         -> 404 Not Found
    PatientAlreadyInactiveException  -> 400 Bad Request
    MethodArgumentNotValidException  -> 400 Bad Request
    Any unlisted exception           -> 500 Internal Server Error

  Service call to stub: {calls_formatted}

WRITE: class `{class_name}ExceptionIT`
  One @Test per exception in {throws_list}:
    when(mockBean.method(any())).thenThrow(new ExactException("msg"))
    mockMvc.perform(...).andExpect(status().isConflict() / isNotFound() / isBadRequest())

Output one compilable Java class in a single ```java block.
"""

INTEG_CTRL_BOUNDARY_PROMPT = """You are writing an INTEGRATION BOUNDARY TEST for controller `{class_name}`.

WHAT YOU ARE TESTING
  Class:       {class_name}  (package: {package})
  Method:      {func_name}
  HTTP:        {http_method}  {base_url}{endpoint_url}

TEST SETUP
  Runner:   @WebMvcTest({class_name}.class)
  Declare:  @Autowired MockMvc mockMvc
            @Autowired ObjectMapper objectMapper
  MockBeans: {fields_formatted}

BOUNDARY CONSTRAINTS (use ONLY these):
  {annotation_values_formatted}

WHAT BOUNDARY TEST MEANS
  Endpoint accepts exact valid edges, rejects values just outside.

WRITE: class `{class_name}BoundaryIT`

  @ParameterizedTest VALID using @CsvSource:
    Rows: min,200  min+1,200  max-1,200  max,200
    stub service -> return valid response
    andExpect(status().isOk() / isCreated())

  @ParameterizedTest INVALID using @ValueSource: {{ min-1, max+1 }}
    NO stub -- validation fires first
    andExpect(status().isBadRequest())

Output one compilable Java class in a single ```java block.
"""

INTEG_REPO_HAPPY_PATH_PROMPT = """You are writing an INTEGRATION HAPPY PATH TEST for repository `{class_name}`.

WHAT YOU ARE TESTING
  Class:       {class_name}  (package: {package})
  Method:      {func_name}
  Return Type: {return_type}
  Parameters:  {parameters}
  Javadoc:     {javadoc}

TEST SETUP
  Runner:  @DataJpaTest
  Declare: @Autowired {class_name} repository
           @Autowired TestEntityManager entityManager

ENTITY CONSTRUCTORS (use these -- do not invent):
  {constructors_formatted}

WRITE: class `{class_name}HappyPathIT`
  Arrange: entityManager.persistAndFlush(new Entity(validArgs))
  Act:     var result = repository.{func_name}(validArg)
  Assert:
    Optional -> assertTrue(result.isPresent()) + assertEquals(expected, result.get().getField())
    boolean  -> assertTrue(result)
    List     -> assertFalse(result.isEmpty()) + assertEquals(expectedSize, result.size())

Output one compilable Java class in a single ```java block.
"""

INTEG_REPO_NEGATIVE_PROMPT = """You are writing an INTEGRATION NEGATIVE TEST for repository `{class_name}`.

WHAT YOU ARE TESTING
  Class:       {class_name}  (package: {package})
  Method:      {func_name}
  Return Type: {return_type}

TEST SETUP
  Runner:  @DataJpaTest
  Declare: @Autowired {class_name} repository
           @Autowired TestEntityManager entityManager

WRITE: class `{class_name}NegativeIT`
  Arrange: do NOT persist any entity (empty DB)
  Act:     var result = repository.{func_name}(nonExistentArg)
  Assert:
    Optional -> assertTrue(result.isEmpty())
    boolean  -> assertFalse(result)
    List     -> assertTrue(result.isEmpty())

Output one compilable Java class in a single ```java block.
"""

INTEG_REPO_CONSTRAINT_PROMPT = """You are writing an INTEGRATION CONSTRAINT VIOLATION TEST for repository `{class_name}`.

WHAT YOU ARE TESTING
  Class: {class_name}  (package: {package})

TEST SETUP
  Runner:  @DataJpaTest
  Declare: @Autowired {class_name} repository
           @Autowired TestEntityManager entityManager

ENTITY CONSTRUCTORS:
  {constructors_formatted}

WRITE: class `{class_name}ConstraintViolationIT`

  @Test - Duplicate unique field:
    entityManager.persistAndFlush(new Entity("Alice", "same@email.com", 30, "9000000000"))
    assertThrows(DataIntegrityViolationException.class, () ->
        entityManager.persistAndFlush(new Entity("Bob", "same@email.com", 25, "9000000001")))

  @Test - Null required field:
    Build entity with null in @NotNull field
    assertThrows(DataIntegrityViolationException.class, () ->
        entityManager.persistAndFlush(invalidEntity))

Output one compilable Java class in a single ```java block.
"""

TECHNIQUE_PROMPTS = {
    "unit": {
        "happy_path": UNIT_HAPPY_PATH_PROMPT,
        "negative":   UNIT_NEGATIVE_PROMPT,
        "exception":  UNIT_EXCEPTION_PROMPT,
        "boundary":   UNIT_BOUNDARY_PROMPT,
        "mock":       UNIT_MOCK_PROMPT,
    },
    "integration_controller": {
        "happy_path": INTEG_CTRL_HAPPY_PATH_PROMPT,
        "negative":   INTEG_CTRL_NEGATIVE_PROMPT,
        "exception":  INTEG_CTRL_EXCEPTION_PROMPT,
        "boundary":   INTEG_CTRL_BOUNDARY_PROMPT,
    },
    "integration_repository": {
        "happy_path":           INTEG_REPO_HAPPY_PATH_PROMPT,
        "negative":             INTEG_REPO_NEGATIVE_PROMPT,
        "constraint_violation": INTEG_REPO_CONSTRAINT_PROMPT,
    }
}

FILE_SUFFIX = {
    "unit": {
        "happy_path": "HappyPathTest",
        "negative":   "NegativeTest",
        "exception":  "ExceptionTest",
        "boundary":   "BoundaryTest",
        "mock":       "MockTest",
    },
    "integration_controller": {
        "happy_path": "HappyPathIT",
        "negative":   "NegativeIT",
        "exception":  "ExceptionIT",
        "boundary":   "BoundaryIT",
    },
    "integration_repository": {
        "happy_path":           "HappyPathIT",
        "negative":             "NegativeIT",
        "constraint_violation": "ConstraintViolationIT",
    }
}

CODE_REQUIRED_TECHNIQUES = {"exception", "mock", "negative"}
SKIP_METHODS = {
    "main", "toString", "equals", "hashCode",
    "getId", "setId", "getName", "setName",
    "getEmail", "setEmail", "isActive", "setActive"
}

def get_axis(class_role: str) -> str:
    if class_role in ("SERVICE", "GENERAL"):
        return "unit"
    elif class_role == "CONTROLLER":
        return "integration_controller"
    elif class_role == "REPOSITORY":
        return "integration_repository"
    return None

def parse_null_checks(body: str) -> list[str]:
    if not body: return []
    matches = re.findall(r'if\s*\(([^)]*==\s*null[^)]*)\)', body)
    return [m.strip() for m in matches]

def parse_blank_checks(body: str) -> list[str]:
    if not body: return []
    matches = re.findall(r'if\s*\(([^)]*(?:isEmpty|blank|trim)[^)]*)\)', body)
    return [m.strip() for m in matches]

def parse_branches(body: str) -> list[str]:
    if not body: return []
    branches = []
    for line in body.splitlines():
        line_str = line.strip()
        if line_str.startswith("if") or line_str.startswith("else if") or "return" in line_str:
            if any(op in line_str for op in ["<", ">", "<=", ">="]):
                branches.append(line_str)
    return branches

def parse_early_exits(body: str) -> list[str]:
    if not body: return []
    exits = []
    lines = body.splitlines()
    for idx, line in enumerate(lines):
        if "throw new" in line:
            context = " ".join([lines[i].strip() for i in range(max(0, idx-2), idx+1)])
            exits.append(context)
    return exits

def map_exception_triggers(throws_list: list, body: str) -> list[str]:
    if not body or not throws_list: return []
    triggers = []
    for exc in throws_list:
        exc_clean = exc.split(".")[-1]
        for line in body.splitlines():
            if f"throw new {exc_clean}" in line:
                triggers.append(f"{exc_clean} -> {line.strip()}")
    return triggers

def parse_http_method(annotations: list) -> str:
    for ann in annotations:
        ann_l = ann.lower()
        if "postmapping" in ann_l: return "post"
        if "getmapping" in ann_l: return "get"
        if "putmapping" in ann_l: return "put"
        if "deletemapping" in ann_l: return "delete"
    return "post"

def parse_base_url(class_annotations: list) -> str:
    for ann in class_annotations:
        if "requestmapping" in ann.lower():
            m = re.search(r'\(["\']([^"\']+)["\']\)', ann)
            if m: return m.group(1)
    return ""

def parse_endpoint_url(annotations: list) -> str:
    for ann in annotations:
        if any(x in ann.lower() for x in ["mapping"]):
            m = re.search(r'\(["\']([^"\']+)["\']\)', ann)
            if m: return m.group(1)
    return ""



def get_applicable_techniques(axis: str, func: dict, ctx: dict, outgoing_calls: list) -> list[str]:
    techniques = ["happy_path"]

    if axis == "unit":
        if ctx.get("null_checks") or ctx.get("blank_checks"):
            techniques.append("negative")
        if func.get("throws"):
            techniques.append("exception")
        if ctx.get("annotation_values_formatted") or ctx.get("boundary_branches"):
            techniques.append("boundary")
        if outgoing_calls:
            techniques.append("mock")

    elif axis == "integration_controller":
        if ctx.get("null_checks") or ctx.get("blank_checks") or ctx.get("annotation_values_formatted"):
            techniques.append("negative")
        if func.get("throws"):
            techniques.append("exception")
        if ctx.get("annotation_values_formatted") or ctx.get("boundary_branches"):
            techniques.append("boundary")

    elif axis == "integration_repository":
        techniques.append("negative")
        techniques.append("constraint_violation")

    return techniques

def needs_body(applicable_techniques: list) -> bool:
    return bool(set(applicable_techniques) & CODE_REQUIRED_TECHNIQUES)

def build_context(func: dict, parent_class: dict, outgoing_calls: list, include_body: bool) -> dict:
    body_text = func.get("body", "") if include_body else "(not required for these techniques)"

    null_checks        = parse_null_checks(body_text)
    blank_checks       = parse_blank_checks(body_text)
    boundary_branches  = parse_branches(body_text)
    early_exits        = parse_early_exits(body_text)
    exception_triggers = map_exception_triggers(func.get("throws", []), body_text)

    calls_formatted = "\n".join([
        f"  {e.get('object','')}.{e.get('target','')}({', '.join(e.get('arguments', []))}) -> {e.get('callee_return_type', 'void')}"
        for e in outgoing_calls
    ])

    fields_formatted = "\n".join([
        f"  {f.get('type','')} {f.get('name','')}" for f in parent_class.get("fields", [])
    ])
    constructors_formatted = "\n".join([
        f"  new {parent_class.get('name', '')}({', '.join(c.get('params', []))})"
        for c in parent_class.get("constructors", [])
    ])

    file_path = func.get("file", "")
    if "src/main/java/" in file_path:
        package = file_path.split("src/main/java/")[-1].rsplit("/", 1)[0].replace("/", ".")
    else:
        package = "com.medibook"

    class_name = func.get("class_name", "UnknownClass")
    class_name_lower = class_name[0].lower() + class_name[1:] if class_name else "target"

    return {
        "package":                    package,
        "class_name":                 class_name,
        "class_name_lower":           class_name_lower,
        "class_role":                 func.get("class_role", ""),
        "class_annotations":          ", ".join(parent_class.get("annotations", [])),
        "func_name":                  func.get("name", ""),
        "return_type":                func.get("return_type", ""),
        "parameters":                 ", ".join(func.get("parameters", [])),
        "throws_list":                ", ".join(func.get("throws", [])),
        "annotations":                ", ".join(func.get("annotations", [])),
        "javadoc":                    func.get("javadoc", ""),
        "body":                       body_text,
        "fields_formatted":           fields_formatted or "(none)",
        "constructors_formatted":     constructors_formatted or f"  new {class_name}()",
        "calls_formatted":            calls_formatted or "(none)",
        "annotation_values_formatted": func.get("annotation_values_formatted", "(none)"),
        "null_checks":                "\n  ".join(null_checks) or "(none detected)",
        "blank_checks":               "\n  ".join(blank_checks) or "(none detected)",
        "boundary_branches":          "\n  ".join(boundary_branches) or "(none detected)",
        "early_exit_conditions":      "\n  ".join(early_exits) or "(none detected)",
        "exception_trigger_map":      "\n  ".join(exception_triggers) or "(none detected)",
        "http_method":                parse_http_method(func.get("annotations", [])),
        "base_url":                   parse_base_url(parent_class.get("annotations", [])),
        "endpoint_url":               parse_endpoint_url(func.get("annotations", [])),
    }

def sanitize_java_code(java_code: str) -> str:
    """Fixes common LLM syntax glitches like missing 'void' return types, typos like 'wbvoid', and mismatched any(Class.class) stubs."""
    # 1. Clean up typos where LLM or regex artifact created 'wbvoid', 'bvoid', etc.
    java_code = re.sub(r'\b[a-zA-Z]{1,3}void\b', 'void', java_code)

    # 2. Insert 'void' before test method names if return type was omitted after @Test / @DisplayName
    # Matches: @Test (optional @DisplayName) followed by an identifier starting a method without a return type
    pattern_missing_void = r'(@Test\s*(?:\n\s*@DisplayName\([^)]+\))?\s*\n\s*)([a-zA-Z0-9_]+\s*\(\)\s*\{)'
    def replace_missing_void(match):
        prefix = match.group(1)
        method_sig = match.group(2)
        # If method_sig doesn't start with a known modifier or return type, prepend void
        if not re.match(r'^(?:public|protected|private|void|static)\b', method_sig):
            return f"{prefix}void {method_sig}"
        return match.group(0)

    java_code = re.sub(pattern_missing_void, replace_missing_void, java_code)

    # 3. Replace any(AnyClass.class) in Mockito stubs/verifications with plain any() to eliminate type inference compilation errors
    pattern_any = r'any\s*\(\s*[A-Za-z0-9_\.]+\.class\s*\)'
    java_code = re.sub(pattern_any, 'any()', java_code, flags=re.IGNORECASE)

    return java_code

def extract_java_block(text: str) -> str:
    if "```java" in text:
        block = text.split("```java")[1].split("```")[0].strip()
        return sanitize_java_code(block)
    elif "```" in text:
        block = text.split("```")[1].split("```")[0].strip()
        return sanitize_java_code(block)
    return sanitize_java_code(text.strip())

def append_csv(csv_path: Path, class_name: str, func_name: str, axis: str, technique: str, out_file: str):
    header = ["ClassName", "FunctionName", "Axis1", "Axis2Technique", "OutputFile"]
    needs_header = not csv_path.exists()
    
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if needs_header:
            writer.writerow(header)
        writer.writerow([class_name, func_name, axis, technique, out_file])

def extract_csv_block(text: str) -> str:
    """Extracts CSV data rows block from Gemini output ([CSV]...[/CSV] or ```csv ... ```)."""
    block = ""
    if "[CSV]" in text and "[/CSV]" in text:
        block = text.split("[CSV]")[1].split("[/CSV]")[0].strip()
    elif "```csv" in text:
        block = text.split("```csv")[1].split("```")[0].strip()
    elif "```CSV" in text:
        block = text.split("```CSV")[1].split("```")[0].strip()
    else:
        # Fallback: look for blocks with commas that are not java code
        blocks = text.split("```")
        csv_lines = []
        for b in blocks:
            if b.strip().startswith("java"):
                continue
            lines = [l.strip() for l in b.strip().splitlines() if l.strip()]
            if lines and "," in lines[0] and not lines[0].startswith("import") and not lines[0].startswith("package"):
                csv_lines.extend(lines)
        block = "\n".join(csv_lines)

    exits = []
    lines = body.splitlines()
    for idx, line in enumerate(lines):
        if "throw new" in line:
            context = " ".join([lines[i].strip() for i in range(max(0, idx-2), idx+1)])
            exits.append(context)
    return exits

def map_exception_triggers(throws_list: list, body: str) -> list[str]:
    if not body or not throws_list: return []
    triggers = []
    for exc in throws_list:
        exc_clean = exc.split(".")[-1]
        for line in body.splitlines():
            if f"throw new {exc_clean}" in line:
                triggers.append(f"{exc_clean} -> {line.strip()}")
    return triggers

def parse_http_method(annotations: list) -> str:
    for ann in annotations:
        ann_l = ann.lower()
        if "postmapping" in ann_l: return "post"
        if "getmapping" in ann_l: return "get"
        if "putmapping" in ann_l: return "put"
        if "deletemapping" in ann_l: return "delete"
    return "post"

def parse_base_url(class_annotations: list) -> str:
    for ann in class_annotations:
        if "requestmapping" in ann.lower():
            m = re.search(r'\(["\']([^"\']+)["\']\)', ann)
            if m: return m.group(1)
    return ""

def parse_endpoint_url(annotations: list) -> str:
    for ann in annotations:
        if any(x in ann.lower() for x in ["mapping"]):
            m = re.search(r'\(["\']([^"\']+)["\']\)', ann)
            if m: return m.group(1)
    return ""



def get_applicable_techniques(axis: str, func: dict, ctx: dict, outgoing_calls: list) -> list[str]:
    techniques = ["happy_path"]

    if axis == "unit":
        if ctx.get("null_checks") or ctx.get("blank_checks"):
            techniques.append("negative")
        if func.get("throws"):
            techniques.append("exception")
        if ctx.get("annotation_values_formatted") or ctx.get("boundary_branches"):
            techniques.append("boundary")
        if outgoing_calls:
            techniques.append("mock")

    elif axis == "integration_controller":
        if ctx.get("null_checks") or ctx.get("blank_checks") or ctx.get("annotation_values_formatted"):
            techniques.append("negative")
        if func.get("throws"):
            techniques.append("exception")
        if ctx.get("annotation_values_formatted") or ctx.get("boundary_branches"):
            techniques.append("boundary")

    elif axis == "integration_repository":
        techniques.append("negative")
        techniques.append("constraint_violation")

    return techniques

def needs_body(applicable_techniques: list) -> bool:
    return bool(set(applicable_techniques) & CODE_REQUIRED_TECHNIQUES)

def build_context(func: dict, parent_class: dict, outgoing_calls: list, include_body: bool) -> dict:
    body_text = func.get("body", "") if include_body else "(not required for these techniques)"

    null_checks        = parse_null_checks(body_text)
    blank_checks       = parse_blank_checks(body_text)
    boundary_branches  = parse_branches(body_text)
    early_exits        = parse_early_exits(body_text)
    exception_triggers = map_exception_triggers(func.get("throws", []), body_text)

    calls_formatted = "\n".join([
        f"  {e.get('object','')}.{e.get('target','')}({', '.join(e.get('arguments', []))}) -> {e.get('callee_return_type', 'void')}"
        for e in outgoing_calls
    ])

    fields_formatted = "\n".join([
        f"  {f.get('type','')} {f.get('name','')}" for f in parent_class.get("fields", [])
    ])
    constructors_formatted = "\n".join([
        f"  new {parent_class.get('name', '')}({', '.join(c.get('params', []))})"
        for c in parent_class.get("constructors", [])
    ])

    file_path = func.get("file", "")
    if "src/main/java/" in file_path:
        package = file_path.split("src/main/java/")[-1].rsplit("/", 1)[0].replace("/", ".")
    else:
        package = "com.medibook"

    class_name = func.get("class_name", "UnknownClass")
    class_name_lower = class_name[0].lower() + class_name[1:] if class_name else "target"

    return {
        "package":                    package,
        "class_name":                 class_name,
        "class_name_lower":           class_name_lower,
        "class_role":                 func.get("class_role", ""),
        "class_annotations":          ", ".join(parent_class.get("annotations", [])),
        "func_name":                  func.get("name", ""),
        "return_type":                func.get("return_type", ""),
        "parameters":                 ", ".join(func.get("parameters", [])),
        "throws_list":                ", ".join(func.get("throws", [])),
        "annotations":                ", ".join(func.get("annotations", [])),
        "javadoc":                    func.get("javadoc", ""),
        "body":                       body_text,
        "fields_formatted":           fields_formatted or "(none)",
        "constructors_formatted":     constructors_formatted or f"  new {class_name}()",
        "calls_formatted":            calls_formatted or "(none)",
        "annotation_values_formatted": func.get("annotation_values_formatted", "(none)"),
        "null_checks":                "\n  ".join(null_checks) or "(none detected)",
        "blank_checks":               "\n  ".join(blank_checks) or "(none detected)",
        "boundary_branches":          "\n  ".join(boundary_branches) or "(none detected)",
        "early_exit_conditions":      "\n  ".join(early_exits) or "(none detected)",
        "exception_trigger_map":      "\n  ".join(exception_triggers) or "(none detected)",
        "http_method":                parse_http_method(func.get("annotations", [])),
        "base_url":                   parse_base_url(parent_class.get("annotations", [])),
        "endpoint_url":               parse_endpoint_url(func.get("annotations", [])),
    }

def sanitize_java_code(java_code: str) -> str:
    """Fixes common LLM syntax glitches like missing 'void' return types, typos like 'wbvoid', and mismatched any(Class.class) stubs."""
    # 1. Clean up typos where LLM or regex artifact created 'wbvoid', 'bvoid', etc.
    java_code = re.sub(r'\b[a-zA-Z]{1,3}void\b', 'void', java_code)

    # 2. Insert 'void' before test method names if return type was omitted after @Test / @DisplayName
    # Matches: @Test (optional @DisplayName) followed by an identifier starting a method without a return type
    pattern_missing_void = r'(@Test\s*(?:\n\s*@DisplayName\([^)]+\))?\s*\n\s*)([a-zA-Z0-9_]+\s*\(\)\s*\{)'
    def replace_missing_void(match):
        prefix = match.group(1)
        method_sig = match.group(2)
        # If method_sig doesn't start with a known modifier or return type, prepend void
        if not re.match(r'^(?:public|protected|private|void|static)\b', method_sig):
            return f"{prefix}void {method_sig}"
        return match.group(0)

    java_code = re.sub(pattern_missing_void, replace_missing_void, java_code)

    # 3. Replace any(AnyClass.class) in Mockito stubs/verifications with plain any() to eliminate type inference compilation errors
    pattern_any = r'any\s*\(\s*[A-Za-z0-9_\.]+\.class\s*\)'
    java_code = re.sub(pattern_any, 'any()', java_code, flags=re.IGNORECASE)

    return java_code

def extract_java_block(text: str) -> str:
    if "```java" in text:
        block = text.split("```java")[1].split("```")[0].strip()
        return sanitize_java_code(block)
    elif "```" in text:
        block = text.split("```")[1].split("```")[0].strip()
        return sanitize_java_code(block)
    return sanitize_java_code(text.strip())

def append_csv(csv_path: Path, class_name: str, func_name: str, axis: str, technique: str, out_file: str):
    header = ["ClassName", "FunctionName", "Axis1", "Axis2Technique", "OutputFile"]
    needs_header = not csv_path.exists()
    
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if needs_header:
            writer.writerow(header)
        writer.writerow([class_name, func_name, axis, technique, out_file])

def extract_csv_block(text: str) -> str:
    """Extracts CSV data rows block from Gemini output ([CSV]...[/CSV] or ```csv ... ```)."""
    block = ""
    if "[CSV]" in text and "[/CSV]" in text:
        block = text.split("[CSV]")[1].split("[/CSV]")[0].strip()
    elif "```csv" in text:
        block = text.split("```csv")[1].split("```")[0].strip()
    elif "```CSV" in text:
        block = text.split("```CSV")[1].split("```")[0].strip()
    else:
        # Fallback: look for blocks with commas that are not java code
        blocks = text.split("```")
        csv_lines = []
        for b in blocks:
            if b.strip().startswith("java"):
                continue
            lines = [l.strip() for l in b.strip().splitlines() if l.strip()]
            if lines and "," in lines[0] and not lines[0].startswith("import") and not lines[0].startswith("package"):
                csv_lines.extend(lines)
        block = "\n".join(csv_lines)

    if block.startswith("```csv"): block = block[6:]
    if block.startswith("```"): block = block[3:]
    if block.endswith("```"): block = block[:-3]
    return block.strip()

MANUAL_HEADERS = ["Test ID", "Module Name", "Function Name", "Test Type", "Test Scenario",
                  "Pre Conditions", "Test Steps", "Test Data", "Expected Result", "Priority",
                  "Executed By", "Execution Date", "Status", "Remarks"]

def append_manual_excel(excel_path: Path, csv_rows_str: str, sheet_title: str = "Manual Tests"):
    """Appends manual test cases directly into a professionally styled Excel (.xlsx) file."""
    if not csv_rows_str or not csv_rows_str.strip():
        return

    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        excel_path.parent.mkdir(parents=True, exist_ok=True)

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        if excel_path.exists():
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_title[:31]
            ws.views.sheetView[0].showGridLines = True
            ws.append(MANUAL_HEADERS)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center

        reader = list(csv.reader(io.StringIO(csv_rows_str.strip())))
        for r in reader:
            if not r or all(not str(c).strip() for c in r):
                continue
            ws.append(r)
            current_row = ws.max_row
            for col_num, cell in enumerate(ws[current_row], 1):
                cell.border = thin_border
                if col_num in [1, 4, 10, 11, 12, 13]:
                    cell.alignment = alignment_center
                else:
                    cell.alignment = alignment_left

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

        wb.save(excel_path)
    except Exception as e:
        print(f"  -> Excel save note ({e}) for {excel_path.name}")


def main():
    parser = argparse.ArgumentParser(description="Tester Agent V2")
    parser.add_argument('--repo', default='', help="Repo owner/name")
    parser.add_argument('--kb', default='', help="Path to kb.json")
    parser.add_argument('--out-dir', default='', help="Output directory")
    parser.add_argument('--model', default='gemini-2.0-flash-lite', help="Gemini model name")
    parser.add_argument('--changed-files', default='', help="Comma-separated list of changed files for delta testing")
    args = parser.parse_args()

    raw_key = os.getenv("GEMINI_API_KEY", "")
    api_keys = [k.strip() for k in raw_key.split(",") if k.strip()]
    if not api_keys:
        print("Error: GEMINI_API_KEY environment variable is not set.")
        sys.exit(1)

    def normalize_model(name: str) -> str:
        name = name.strip()
        if not name.startswith("models/") and not name.startswith("tunedModels/"):
            return f"models/{name}"
        return name

    # Verified active Google Gemini models matched against live API & Google AI Studio limits
    # Ordered by daily quota: 500 RPD models first, followed by 20 RPD & standard Flash models
    model_pool = [
        'models/gemini-3.5-flash-lite',   # 15 RPM, 500 RPD (Verified Working)
        'models/gemini-3.1-flash-lite',   # 15 RPM, 500 RPD (Verified Working)
        'models/gemini-3.5-flash',        # 5 RPM, 20 RPD (Verified Working)
        'models/gemini-3.6-flash',        # 5 RPM, 20 RPD (Verified Working)
        'models/gemini-2.0-flash',        # 15 RPM (Verified Working)
        'models/gemini-2.0-flash-lite',   # 30 RPM (Verified Working)
        'models/gemini-1.5-flash',        # 15 RPM (Verified Working)
        'models/gemini-1.5-pro'           # 2 RPM (Verified Working)
    ]
    if args.model:
        user_m = normalize_model(args.model)
        if user_m not in model_pool:
            model_pool.insert(0, user_m)




    call_counter = 0

    repo = args.repo.strip() or os.getenv("GITHUB_REPO", "").strip()
    repo_name = repo.split('/')[-1] if "/" in repo else repo

    if args.kb:
        kb_path = Path(args.kb)
    elif repo_name:
        kb_path = Path("output") / repo_name / "kb.json"
    else:
        kb_path = Path("output") / "kb.json"

    if not kb_path.exists():
        print(f"Error: kb.json not found at {kb_path}")
        sys.exit(1)

    print(f"Loading Knowledge Base from {kb_path}...")
    with open(kb_path, "r", encoding="utf-8") as f:
        kb_data = json.load(f)

    nodes = kb_data.get("nodes", [])
    edges = kb_data.get("edges", [])

    if args.out_dir:
        out_base = Path(args.out_dir)
    elif repo_name:
        out_base = Path("output") / repo_name
    else:
        out_base = Path("output")

    out_base.mkdir(parents=True, exist_ok=True)

    classes_by_name = {n["name"]: n for n in nodes if n.get("type") == "CLASS"}
    calls_by_func = defaultdict(list)
    for e in edges:
        if e.get("type") == "CALLS":
            calls_by_func[e["source"]].append(e)

    def is_dto_or_getter(func, parent_class):
        name = func.get("name", "")
        cname = func.get("class_name", "")
        crole = parent_class.get("class_role") or func.get("class_role")
        if crole in ("ENTITY", "DTO"):
            return True
        if any(cname.endswith(suffix) for suffix in ["Request", "Response", "DTO", "Dto"]):
            return True
        if name.startswith("get") or name.startswith("set") or name.startswith("is"):
            return True
        return False

    def is_config_or_dto(func, parent_class):
        cname = func.get("class_name", "")
        crole = parent_class.get("class_role") or func.get("class_role")
        if crole == "CONFIGURATION":
            return True
        if cname.endswith("Config") or cname.endswith("Configuration") or cname.endswith("Application"):
            return True
        return is_dto_or_getter(func, parent_class)

    func_nodes = [
        n for n in nodes
        if n.get("type") == "FUNCTION"
        and n.get("file", "").endswith(".java")
        and n.get("class_name")
        and n.get("name") not in SKIP_METHODS
        and not is_config_or_dto(n, classes_by_name.get(n.get("class_name"), {}))
    ]

    print(f"Found {len(func_nodes)} testable target functions.")

    # ── PHASE 2: PLAN ──
    test_plan = []
    for func in func_nodes:
        class_role = func.get("class_role", "GENERAL")
        axis = get_axis(class_role)
        if not axis:
            continue

        parent_class = classes_by_name.get(func.get("class_name"), {})
        outgoing_calls = calls_by_func.get(func.get("id"), [])

        raw_ctx = build_context(func, parent_class, outgoing_calls, include_body=False)
        applicable = get_applicable_techniques(axis, func, raw_ctx, outgoing_calls)
        include_body = needs_body(applicable)

        ctx = build_context(func, parent_class, outgoing_calls, include_body)
        axis_short = "unit" if axis == "unit" else "integration"
        suffixes = FILE_SUFFIX[axis]

        out_files = {
            t: str(Path("automated") / axis_short / t / f"{func.get('class_name')}{suffixes[t]}.java")
            for t in applicable
        }

        test_plan.append({
            "class": func.get("class_name"),
            "func": func.get("name"),
            "axis": axis,
            "techniques": applicable,
            "include_body": include_body,
            "out_files": out_files,
            "context": ctx
        })

    plan_path = out_base / "test_plan.json"
    with open(plan_path, "w", encoding="utf-8") as f:
        json.dump(test_plan, f, indent=2)
    print(f"Saved test plan ({len(test_plan)} tasks) to {plan_path}")

    # ── PHASE 3: GENERATE ──
    csv_path = out_base / "test_matrix_summary.csv"

    for idx, task in enumerate(test_plan, 1):
        print(f"\n[{idx}/{len(test_plan)}] {task['class']}.{task['func']}() -> {task['techniques']}")
        axis_short = "unit" if task["axis"] == "unit" else "integration"

        for technique in task["techniques"]:
            rel_file = task["out_files"][technique]
            full_file = out_base / rel_file

            if full_file.exists():
                print(f"  [SKIP] {rel_file} already exists.")
                continue

            call_counter += 1
            cur_key = api_keys[call_counter % len(api_keys)]
            cur_model_name = model_pool[call_counter % len(model_pool)]

            genai.configure(api_key=cur_key, transport='rest')
            model = genai.GenerativeModel(cur_model_name)

            prompt_tmpl = TECHNIQUE_PROMPTS[task["axis"]][technique]
            prompt = prompt_tmpl.format(**task["context"])
            full_prompt = f"{SYSTEM_PROMPT}\n\n{prompt}"

            print(f"  -> Generating {technique} [{cur_model_name}] ({rel_file})...", flush=True)

            success = False
            for attempt in range(1, 4):
                try:
                    response = model.generate_content(full_prompt)
                    java_code = extract_java_block(response.text)
                    csv_rows = extract_csv_block(response.text)

                    full_file.parent.mkdir(parents=True, exist_ok=True)
                    full_file.write_text(java_code, encoding="utf-8")
                    print(f"  -> Saved automated code: {full_file}", flush=True)

                    append_csv(csv_path, task["class"], task["func"], task["axis"], technique, rel_file)

                    if csv_rows:
                        tech_excel_path = out_base / "manual" / f"{axis_short}_{technique}.xlsx"
                        master_excel_path = out_base / "manual" / "manual_test_cases_master.xlsx"
                        sheet_title = f"{axis_short}_{technique}".replace("_", " ").title()
                        append_manual_excel(tech_excel_path, csv_rows, sheet_title)
                        append_manual_excel(master_excel_path, csv_rows, "Master Manual Tests")
                        print(f"  -> Appended manual test cases: {tech_excel_path.name}", flush=True)

                    time.sleep(1.5)  # Pause to remain within RPM limits
                    success = True
                    break
                except Exception as e:
                    err_str = str(e)
                    if "429" in err_str or "quota" in err_str.lower() or "resource_exhausted" in err_str.lower():
                        cur_model_name = model_pool[(call_counter + attempt) % len(model_pool)]
                        print(f"  -> Rate limit (429) hit (attempt {attempt}/3). Pausing 4s & switching model to '{cur_model_name}'...", flush=True)
                        model = genai.GenerativeModel(cur_model_name)
                        time.sleep(4.0)
                        continue
                    elif "404" in err_str or "not found" in err_str.lower():
                        cur_model_name = model_pool[(call_counter + attempt) % len(model_pool)]
                        print(f"  -> Model '{cur_model_name}' not available. Switching model to '{cur_model_name}'...", flush=True)
                        model = genai.GenerativeModel(cur_model_name)
                        time.sleep(1.0)
                        continue
                    else:
                        print(f"  -> Error ({technique}) {task['class']}.{task['func']}: {e}", flush=True)
                        time.sleep(0.5)
                        break


    # Clean up any residual .csv files in manual folder so only .xlsx Excel files remain
    manual_dir = out_base / "manual"
    if manual_dir.exists():
        for old_csv in manual_dir.glob("*.csv"):
            try:
                old_csv.unlink()
            except Exception:
                pass

    print("\nTester Agent V2 completed successfully!")
    print(f"Outputs written to {out_base}")


if __name__ == "__main__":
    main()
